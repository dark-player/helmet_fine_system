"""
Creates sample bike_owners.xlsx with demo data
Run this once to generate the Excel file
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bike Owners"

# Headers
headers = ["Number Plate", "Owner Name", "Phone", "Email", "Address", "City", "State"]
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Sample data - demo bike owners
data = [
    ["MH12AB1234", "Rahul Sharma",   "9876543210", "rahul.sharma@gmail.com",   "123 MG Road",       "Pune",    "Maharashtra"],
    ["MH14CD5678", "Priya Patil",    "9823456789", "priya.patil@gmail.com",    "45 Shivaji Nagar",  "Pune",    "Maharashtra"],
    ["MH01EF9012", "Amit Kumar",     "9712345678", "amit.kumar@gmail.com",     "78 Baner Road",     "Pune",    "Maharashtra"],
    ["MH15GH3456", "Suresh Jadhav",  "9645678901", "suresh.jadhav@gmail.com",  "12 Kothrud",        "Pune",    "Maharashtra"],
    ["MH20IJ7890", "Pooja Desai",    "9534567890", "pooja.desai@gmail.com",    "56 Wakad",          "Pune",    "Maharashtra"],
    ["MH11KL2345", "Vijay More",     "9423456789", "vijay.more@gmail.com",     "90 Hadapsar",       "Pune",    "Maharashtra"],
    ["DL01MN6789", "Rohit Singh",    "9312345678", "rohit.singh@gmail.com",    "34 Connaught Place","Delhi",   "Delhi"],
    ["KA03OP1234", "Deepa Nair",     "9201234567", "deepa.nair@gmail.com",     "67 Koramangala",    "Bangalore","Karnataka"],
    ["GJ05QR5678", "Ravi Patel",     "9109876543", "ravi.patel@gmail.com",     "23 CG Road",        "Ahmedabad","Gujarat"],
    ["TN07ST9012", "Meena Iyer",     "9898765432", "meena.iyer@gmail.com",     "11 Anna Nagar",     "Chennai", "Tamil Nadu"],
]

for row_data in data:
    ws.append(row_data)

# Column widths
widths = [15, 18, 14, 30, 25, 15, 15]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, col).column_letter].width = w

wb.save("bike_owners.xlsx")
print("bike_owners.xlsx created successfully!")
print(f"Total owners: {len(data)}")
